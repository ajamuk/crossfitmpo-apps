#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Detector precoz de bajas — CrossFit MPO Parla
=============================================

Para cada socio ACTIVO a fecha de corte estima la probabilidad (%) de causar
baja durante el mes siguiente y explica, en lenguaje natural, los factores de
riesgo concretos de ese socio.

ENFOQUE: scorecard explicable (estilo credit-scoring)
-----------------------------------------------------
Se eligió un modelo de puntuación aditivo en lugar de una "caja negra" porque
el objetivo es ACCIONAR sobre cada socio: cada factor pesa en su dirección
lógica y la MISMA fórmula que calcula el % es la que genera la explicación.

  1) Se comparan 123 bajas del último año (su estado en el momento de irse) con
     251 socios activos (su estado hoy). De ahí se aprende, para cada señal por
     separado, cuánto pesa en el riesgo de baja (regresión logística univariante
     -> el signo de cada factor está garantizado y es interpretable).
  2) Señales del scorecard (todas: "más alto = más riesgo"):
        - Recencia: días desde la última reserva.
        - Poca actividad reciente: reservas en los últimos 30 días (a menos,
          más riesgo).  ← señal dominante en estos datos.
        - Poco historial acumulado: reservas totales (socios con poca "raíz"
          en el club se van más).
  3) Las tres señales se combinan en un score y se convierte a probabilidad.
     El % se RE-CALIBRA a la tasa real de baja mensual del centro (~4%) con
     corrección de prevalencia (King & Zeng, 2001), para que sea interpretable
     como "riesgo de baja el mes que viene".

Nota empírica honesta: en estos datos el "declive relativo" (caída de actividad
respecto al nivel habitual del propio socio) NO resultó predictivo, porque
muchas bajas son fin de contrato con el socio aún activo (mudanza, lesión,
precio). Por eso el motor se apoya en la actividad reciente ABSOLUTA y la
recencia. El declive se conserva como columna informativa, no como peso.

Uso:
    python3 detector_bajas.py
Genera: salida/deteccion_bajas_parla.xlsx
"""
import os
import numpy as np
import pandas as pd
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import StratifiedKFold
from sklearn.metrics import roc_auc_score

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "datos")
OUT_DIR = os.path.join(HERE, "salida")
F_ACT = os.path.join(DATA, "activos_parla_202606.xlsx")
F_BAJ = os.path.join(DATA, "bajas_parla_202506_202604.xlsx")

# --- Parámetro de negocio: tasa de baja mensual real del centro ---------------
# 123 bajas en una ventana de 10 meses sobre una base media ~310 socios -> ~4%.
TASA_BAJA_MENSUAL = 0.04

# Features del scorecard, ya orientadas a "mayor valor = mayor riesgo".
SCORE_FEATURES = ["recencia", "poco_vol30", "poco_hist"]
FEAT_LABEL = {
    "recencia":   "días desde la última reserva",
    "poco_vol30": "reservas en los últimos 30 días (pocas)",
    "poco_hist":  "reservas totales acumuladas (pocas)",
}


def num(s):
    """Convierte a numérico tratando '[SIN DATO]' como NaN."""
    return pd.to_numeric(pd.Series(s).replace("[SIN DATO]", np.nan), errors="coerce")


def build_features(path):
    """Extrae métricas por socio de 'Tabla Maestra' + tendencia mensual."""
    tm = pd.read_excel(path, sheet_name="Tabla Maestra")
    rm = pd.read_excel(path, sheet_name="Reservas Mensuales")
    monthcols = [c for c in rm.columns
                 if isinstance(c, str) and len(c) == 7 and c[4] == "-"]
    rm_idx = rm.set_index("ID")

    rows = []
    for _, r in tm.iterrows():
        ID = r["ID"]
        g = lambda col: num([r[col]]).iloc[0]

        # tendencia desde la rejilla mensual (sólo informativa)
        recent3 = baseline = np.nan
        if ID in rm_idx.index:
            srow = rm_idx.loc[ID]
            if isinstance(srow, pd.DataFrame):
                srow = srow.iloc[0]
            ser = pd.to_numeric(srow[monthcols], errors="coerce")
            nz = ser[ser.notna()]
            if len(nz) > 0:
                last_pos = max(monthcols.index(m) for m in nz.index)
                hist = ser.iloc[:last_pos + 1].fillna(0)
                recent3 = hist.iloc[-3:].mean()
                base_part = hist.iloc[:-3]
                baseline = (base_part[base_part > 0].mean()
                            if (base_part > 0).any() else recent3)

        rows.append(dict(
            ID=ID, Nombre=r["Nombre"], Programa=r.get("Programa", ""),
            Alta=r.get("Alta", ""),
            recency=g("Días última reserva → corte"),
            act30=g("Reservas últimos 30d"),
            act60=g("Reservas últimos 60d"),
            act90=g("Reservas últimos 90d"),
            freq=g("Frecuencia media (res/sem)"),
            antig=g("Antigüedad (meses)"),
            restot=g("Reservas totales"),
            gasto=g("Total gastado (€)"),
            ult_reserva=r.get("Última reserva", ""),
            recent3=recent3, baseline=baseline,
        ))
    d = pd.DataFrame(rows)
    d["decline"] = d["recent3"] / (d["baseline"] + 1e-6)   # informativo
    # --- features orientadas (mayor = más riesgo) ---
    d["recencia"]   = np.log1p(d["recency"].clip(lower=0))
    d["poco_vol30"] = -d["act30"]
    d["poco_hist"]  = -np.log1p(d["restot"].clip(lower=0))
    return d


def fit_scorecard(Z, y):
    """Pesos univariantes (signo garantizado) por feature."""
    w = {}
    for c in SCORE_FEATURES:
        w[c] = LogisticRegression(max_iter=2000).fit(Z[[c]], y).coef_[0][0]
    return np.array([w[c] for c in SCORE_FEATURES]), w


def main():
    da = build_features(F_ACT); da["y"] = 0
    db = build_features(F_BAJ); db["y"] = 1
    full = pd.concat([da, db], ignore_index=True)
    y = full["y"].values

    X = full[SCORE_FEATURES].copy()
    med = X.median()
    X = X.fillna(med)
    mu = X.mean()
    sd = X.std().replace(0, 1)
    Z = (X - mu) / sd

    # --- validación honesta (out-of-fold), recalculando pesos por fold -------
    skf = StratifiedKFold(5, shuffle=True, random_state=1)
    oof = np.zeros(len(y))
    for tr, te in skf.split(Z, y):
        ww, _ = fit_scorecard(Z.iloc[tr], y[tr])
        s_tr = Z.iloc[tr].values @ ww
        s_te = Z.iloc[te].values @ ww
        cal = LogisticRegression(max_iter=2000).fit(s_tr.reshape(-1, 1), y[tr])
        oof[te] = cal.predict_proba(s_te.reshape(-1, 1))[:, 1]
    auc = roc_auc_score(y, oof)

    # --- modelo final --------------------------------------------------------
    weights, wdict = fit_scorecard(Z, y)
    score = Z.values @ weights
    cal = LogisticRegression(max_iter=2000).fit(score.reshape(-1, 1), y)
    a = cal.coef_[0][0]; b0 = cal.intercept_[0]

    # corrección de prevalencia (King & Zeng) sobre el intercepto
    tau = TASA_BAJA_MENSUAL
    ybar = y.mean()
    b0_corr = b0 - np.log((1 - tau) / tau * ybar / (1 - ybar))

    # --- puntuar SOLO activos ------------------------------------------------
    act = da.copy().reset_index(drop=True)
    Za = (act[SCORE_FEATURES].fillna(med) - mu) / sd
    # contribución de cada feature al score (en unidades de log-odds finales)
    contrib = Za.values * weights * a          # (n, k)
    sc_act = Za.values @ weights
    logit = a * sc_act + b0_corr
    act["prob"] = 1 / (1 + np.exp(-logit))

    def tramo(p):
        if p >= 0.15: return "🔴 Crítico"
        if p >= 0.08: return "🟠 Alto"
        if p >= 0.04: return "🟡 Medio"
        return "🟢 Bajo"
    act["Riesgo"] = act["prob"].apply(tramo)

    # --- explicación fiel (mismas contribuciones que el score) ---------------
    def frase(f, v):
        if f == "recencia":
            d = v["recency"]
            if pd.isna(d):
                return None
            return "reservó hoy mismo" if int(d) == 0 else \
                   f"lleva {int(d)} día sin reservar" if int(d) == 1 else \
                   f"lleva {int(d)} días sin reservar"
        if f == "poco_vol30":
            a30 = v["act30"]
            if pd.isna(a30):
                return None
            a30 = int(a30)
            return "ninguna reserva en los últimos 30 días" if a30 == 0 else \
                   "sólo 1 reserva en los últimos 30 días" if a30 == 1 else \
                   f"sólo {a30} reservas en los últimos 30 días"
        if f == "poco_hist":
            rt = v["restot"]
            if pd.isna(rt):
                return None
            rt = int(rt)
            return f"historial corto en el club ({rt} reserva total)" if rt == 1 else \
                   f"historial corto en el club ({rt} reservas totales)"
        return None

    def explicar(i):
        c = contrib[i]
        order = np.argsort(c)[::-1]
        out = []
        for j in order:
            if c[j] <= 0.03:
                break
            fr = frase(SCORE_FEATURES[j], act.iloc[i])
            if fr:
                out.append(fr)
        # contexto adicional siempre útil: declive si es marcado
        v = act.iloc[i]
        if pd.notna(v["decline"]) and v["decline"] < 0.6 and pd.notna(v["baseline"]) and v["baseline"] >= 2:
            out.append(f"su actividad reciente cayó al {v['decline']*100:.0f}% de su nivel habitual")
        if not out:
            return "Perfil de socio comprometido (actividad reciente y recencia saludables)."
        return "; ".join(out[:3]).capitalize() + "."

    act["Motivo_principal"] = [explicar(i) for i in range(len(act))]

    def protector(i):
        c = contrib[i]; j = int(np.argmin(c))
        if c[j] >= -0.03:
            return ""
        v = act.iloc[i]; f = SCORE_FEATURES[j]
        if f == "recencia" and pd.notna(v["recency"]):
            return f"reservó hace muy poco ({int(v['recency'])} días)"
        if f == "poco_vol30" and pd.notna(v["act30"]):
            return f"buen volumen reciente ({int(v['act30'])} reservas/30d)"
        if f == "poco_hist" and pd.notna(v["restot"]):
            return f"socio con mucho recorrido ({int(v['restot'])} reservas)"
        return ""
    act["Factor_protector"] = [protector(i) for i in range(len(act))]

    act = act.sort_values("prob", ascending=False).reset_index(drop=True)

    # --- consola -------------------------------------------------------------
    print(f"AUC (validación cruzada 5-fold): {auc:.3f}")
    print(f"Calibración a tasa baja mensual: {tau*100:.1f}%")
    print(f"Prob. media activos tras calibrar: {act['prob'].mean()*100:.2f}%")
    print("Tramos:\n" + act["Riesgo"].value_counts().to_string())
    print("\nPesos del scorecard (univariantes, estandarizados):")
    for f in SCORE_FEATURES:
        print(f"  {f:11s} {wdict[f]:+.3f}  ({FEAT_LABEL[f]})")
    print("\nTop 10 socios en riesgo:")
    print(act[["Nombre", "prob", "Riesgo", "Motivo_principal"]].head(10).to_string(index=False))

    _write_excel(act, auc, wdict, tau, mu, sd)
    return act


def _write_excel(act, auc, wdict, tau, mu, sd):
    os.makedirs(OUT_DIR, exist_ok=True)
    out_path = os.path.join(OUT_DIR, "deteccion_bajas_parla.xlsx")

    main = pd.DataFrame({
        "Nombre": act["Nombre"],
        "ID": act["ID"],
        "Programa": act["Programa"],
        "% Prob. baja (mes siguiente)": (act["prob"] * 100).round(1),
        "Nivel de riesgo": act["Riesgo"],
        "Motivo principal": act["Motivo_principal"],
        "Factor protector": act["Factor_protector"],
        "Días desde última reserva": act["recency"],
        "Reservas últ. 30d": act["act30"],
        "Reservas últ. 60d": act["act60"],
        "Reservas últ. 90d": act["act90"],
        "Frecuencia hist. (ses/sem)": act["freq"],
        "Actividad reciente vs habitual (%)": (act["decline"] * 100).round(0),
        "Reservas totales": act["restot"],
        "Antigüedad (meses)": act["antig"],
        "Última reserva": act["ult_reserva"],
        "Total gastado (€)": act["gasto"],
    })

    metodologia = pd.DataFrame({"Detector precoz de bajas — CrossFit MPO Parla": [
        "",
        "QUÉ ES",
        "Probabilidad de que cada socio activo cause baja durante el mes siguiente,",
        "con el motivo de riesgo de cada uno para poder priorizar la retención.",
        "",
        "CÓMO SE CALCULA (scorecard explicable)",
        "1) Se compara el comportamiento de 123 bajas del último año (en el momento de",
        "   irse) con el de 251 socios activos (hoy).",
        "2) Tres señales suman riesgo, cada una en su dirección lógica:",
        "      • Recencia: días desde la última reserva (más días = más riesgo).",
        "      • Poca actividad reciente: reservas en 30 días (menos = más riesgo). [dominante]",
        "      • Poco historial acumulado: reservas totales (menos raíz = más riesgo).",
        "3) Se combinan en una puntuación y se transforma en probabilidad, re-calibrada",
        f"   a la tasa real de baja mensual del centro ({tau*100:.0f}%) con corrección de",
        "   prevalencia (King & Zeng, 2001). Así el % es 'riesgo del mes que viene'.",
        "",
        "CALIDAD DEL MODELO",
        f"AUC en validación cruzada (5-fold): {auc:.3f}  (0.5=azar, 1.0=perfecto).",
        "El motor ORDENA bien el riesgo; el % es una estimación calibrada, no una certeza.",
        "",
        "TRAMOS DE RIESGO",
        "🔴 Crítico: ≥ 15%     🟠 Alto: 8–15%     🟡 Medio: 4–8%     🟢 Bajo: < 4%",
        "",
        "CÓMO USARLO",
        "Contactar primero a los socios en rojo y naranja. El 'Motivo principal' dice",
        "sobre qué actuar (recuperar a quien lleva semanas sin venir, reenganchar a",
        "quien casi no reserva, acompañar a los socios nuevos con poco arraigo).",
        "",
        "NOTA EMPÍRICA",
        "Muchas bajas son fin de contrato con el socio aún activo (mudanza, lesión,",
        "precio): esos casos son menos predecibles por uso. El 'declive vs su nivel",
        "habitual' no resultó predictivo en estos datos y se muestra sólo como contexto.",
        "Recalcular cada mes con datos frescos.",
    ]})

    pesos = pd.DataFrame({
        "Señal": SCORE_FEATURES,
        "Descripción": [FEAT_LABEL[f] for f in SCORE_FEATURES],
        "Peso (estandarizado)": [round(wdict[f], 4) for f in SCORE_FEATURES],
        "Media poblacional": [round(mu[f], 4) for f in SCORE_FEATURES],
        "Desv. típica": [round(sd[f], 4) for f in SCORE_FEATURES],
    })

    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        main.to_excel(xw, sheet_name="Detector Bajas", index=False)
        metodologia.to_excel(xw, sheet_name="Metodología", index=False)
        pesos.to_excel(xw, sheet_name="Modelo (pesos)", index=False)
        _format(xw, main)
    print(f"\n✅ Generado: {out_path}")


def _format(xw, main):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    ws = xw.sheets["Detector Bajas"]
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill; c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = [26, 9, 10, 14, 13, 58, 32, 12, 11, 11, 11, 14, 16, 12, 13, 13, 13]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    fills = {"🔴": PatternFill("solid", fgColor="FCA5A5"),
             "🟠": PatternFill("solid", fgColor="FDBA74"),
             "🟡": PatternFill("solid", fgColor="FDE68A"),
             "🟢": PatternFill("solid", fgColor="BBF7D0")}
    risk_col = list(main.columns).index("Nivel de riesgo") + 1
    prob_col = list(main.columns).index("% Prob. baja (mes siguiente)") + 1
    for row in range(2, ws.max_row + 1):
        val = str(ws.cell(row=row, column=risk_col).value or "")
        key = val[0] if val else ""
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).border = border
        if key in fills:
            ws.cell(row=row, column=risk_col).fill = fills[key]
            ws.cell(row=row, column=prob_col).fill = fills[key]
        ws.cell(row=row, column=prob_col).font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


if __name__ == "__main__":
    main()
